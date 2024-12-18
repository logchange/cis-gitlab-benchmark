import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

from src.export.exporter import Exporter


def _save_temp_excel(df, path):
    df.to_excel(path, index=False, engine='openpyxl')


def _add_info_header(ws):
    info_text = (
        "This file was generated using: https://github.com/logchange/cis-gitlab-benchmark\n"
        "Check out CIS GitLab Benchmark implementation advices\n"
        "here: https://github.com/logchange/cis-gitlab-benchmark/blob/main/cis_gitlab_benchmark_v1_0_1_implmentation.md\n"
        "Visit https://github.com/logchange/cis-gitlab-benchmark and leave a star 🌟\n"
        "This file is the result of CIS-Controls assessment\n"
        "CIS GitLab Benchmark v1.0.1 - 04-19-2024"
    )

    # Merging cells A1 to T1 for header row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    cell = ws.cell(row=1, column=1)
    cell.value = info_text

    # Applying styles to the header cell
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    ws.row_dimensions[1].height = 100  # Slightly increase row height
    ws.column_dimensions['A'].width = 40


def _structure_data(data):
    structured_data = {}
    for item in data:
        name_with_url = f"{item['name']} ({item['url']})"
        structured_data[name_with_url] = {
            control.to_dict()['control_name']: (control.to_dict()['passed'], control.to_dict()['more_info'])
            for control in item['controls_result']
        }
    return structured_data


def _prepare_dataframe(data):
    formatted_data = _structure_data(data)
    df = pd.DataFrame.from_dict(formatted_data, orient='index')
    columns_passed = {col: [val[0] for val in df[col].values] for col in df.columns}
    return pd.DataFrame(columns_passed, index=df.index).reset_index().rename(columns={'index': 'name (url)'})


class XlsxExporter(Exporter):
    TRUE_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    FALSE_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    NEUTRAL_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    THIN_BORDER = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    def export(self, data, group_name: str):
        df = _prepare_dataframe(data)
        group_name = group_name.replace("/", "-")
        temp_path = f"cis-gitlab-benchmark-{group_name}.xlsx"
        _save_temp_excel(df, temp_path)
        self._format_excel(temp_path, data)

    def _format_excel(self, path, data):
        wb = load_workbook(path)
        ws = wb.active
        ws.move_range("A1:Z1000", rows=1, cols=0)

        _add_info_header(ws)

        formatted_data = _structure_data(data)
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            for j, cell in enumerate(row, start=1):
                self._apply_formatting(i, j, cell, formatted_data, ws)

        wb.save(path)

    def _apply_formatting(self, row_idx, col_idx, cell, formatted_data, ws):
        if row_idx == 2:  # Adjusted to account for new info header at the top
            self._format_header(cell)
        elif col_idx == 1:
            cell.fill = self.NEUTRAL_FILL
            cell.border = self.THIN_BORDER
        else:
            self._format_passed_value(cell, formatted_data, ws, row_idx, col_idx)

    def _format_header(self, cell):
        cell.fill = self.NEUTRAL_FILL
        cell.alignment = Alignment(textRotation=60)

    def _format_passed_value(self, cell, formatted_data, ws, row_idx, col_idx):
        value = cell.value
        if value is True:
            cell.fill = self.TRUE_FILL
        elif value is False:
            cell.fill = self.FALSE_FILL

        # Retrieve keys safely
        row_key = ws.cell(row=row_idx, column=1).value
        col_key = ws.cell(row=2, column=col_idx).value

        # Check if both row_key and col_key exist in the dictionary before adding a comment
        if row_key in formatted_data and col_key in formatted_data[row_key]:
            more_info = formatted_data[row_key][col_key][1]
            if more_info:
                cell.comment = Comment(more_info, "logchange.dev")
